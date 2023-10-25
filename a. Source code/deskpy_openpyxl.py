import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

class Excel():
    def new_book(self, p1, fw, p2):
        self.to_scan = p1.text()
        self.final_workbook = fw.text()
        self.to_save = p2.text()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['#', 'ID', 'CLIENTE', 'PAGARÉ', 'OTROS DOCUMENTOS', 'INFORMACIÓN GENERAL', 'APROBACIONES CREDITICIAS', 'INF. ANÁLISIS CAPACIDAD', 'RESULTADOS DE ANÁLISIS', 'DOCUMENTOS FALTANTES', 'DOCUMENTOS EN CARPETA'])
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 18
        worksheet.column_dimensions['C'].width = 45
        worksheet.column_dimensions['D'].width = 18
        worksheet.column_dimensions['E'].width = 40
        worksheet.column_dimensions['F'].width = 40
        worksheet.column_dimensions['G'].width = 40
        worksheet.column_dimensions['H'].width = 40
        worksheet.column_dimensions['I'].width = 40
        worksheet.column_dimensions['J'].width = 40
        worksheet.column_dimensions['K'].width = 40
        worksheet.row_dimensions[1].height = 30
        self.wb_border = Side(border_style='thin')
        for th in worksheet['A1:K1']:
            for cell in th:
                cell.font = Font(name='Segoe UI', size=12, color='FFFFFF', bold=True)
                cell.fill = PatternFill(start_color="113A44", end_color="113A44", fill_type='solid')
                cell.border = Border(top=self.wb_border, right=self.wb_border, bottom=self.wb_border, left=self.wb_border)
                cell.alignment = Alignment(horizontal='center')
        self.wb_tdcell = PatternFill(start_color='E0EBEC', end_color='E0EBEC', fill_type='solid')
        filters = worksheet.auto_filter
        filters.ref = 'A1:K1'
        worksheet.freeze_panes = 'A2'
        workbook.save(f'{self.to_save}/{self.final_workbook}.xlsx')

    def get_tree(self):
        self.tree = os.listdir(self.to_scan)
        self.iterator_tree = iter(self.tree)
        self.next_folder = next(self.iterator_tree)

    def sck_folder(self):
        myexcel = f'{self.to_save}{self.final_workbook}.xlsx'
        wb = openpyxl.load_workbook(myexcel)
        ws = wb.active
        ltree = len(self.tree)
        self.start_row = ws.max_row
        self.main_folder = []
        for leaf in self.tree:
            self.main_folder.append(f'{self.to_scan}{self.next_folder}/')
            try: self.next_folder = next(self.iterator_tree)
            except: pass
        self.intermitent_highlight = PatternFill(start_color="CFF9EC", end_color="CFF9EC", fill_type='solid')
        self.highlight = False
        for sub_folder in self.main_folder:
            data_recovery = sub_folder.split('/')
            data_recovery = data_recovery[-2]
            data_recovery = data_recovery.split(' ')
            try: column_d = data_recovery.pop()
            except: pass
            try:
                column_c = data_recovery[1:]
                column_c = ' '.join(column_c)
            except: column_c = ''
            try:
                column_b = data_recovery[0]
            except:
                data_recovery = sub_folder.split('/')
                column_b = data_recovery[-2]
            if column_d == column_b: column_d = 'Sin datos'
            if column_c == '': column_c = 'Sin datos'
            self.start_row = ws.max_row + 1
            numeral = self.main_folder.index(sub_folder) + 1
            if numeral < 10: numeral = f'00000{numeral}'
            elif numeral > 9 and numeral < 100: numeral = f'0000{numeral}'
            elif numeral > 99 and numeral < 1000: numeral = f'000{numeral}'
            elif numeral > 999 and numeral < 10000: numeral = f'00{numeral}'
            else: numeral = f'0{numeral}'
            os.system('cls')
            print(f'Procesando expediente {numeral} de {ltree} / {sub_folder}')
            ws[f'A{self.start_row}'].value = numeral
            ftitle = sub_folder.split('/')
            ftitle = ftitle[-2]
            self.collect_all_existing_files = []
            self.subf_dataset = []
            try:
                self.folder_zr = os.listdir(f'{sub_folder}0. OTROS DOCUMENTOS')
                col_e = self.start_row
                if len(os.listdir(f'{sub_folder}0. OTROS DOCUMENTOS')) > 0:
                    print('\t0. OTROS DOCUMENTOS')
                    for file in self.folder_zr:
                        file = file.replace(ftitle,'').replace(' .pdf','.pdf').replace(' .png','.png').replace(' .jpg','.jpg').replace(' .jpeg','.jpeg').replace(' .webm','.webm').replace(' .mp4','.mp4')
                        print(f'\t\t>>> {file}')
                        self.subf_dataset.append(file)
                        self.collect_all_existing_files.append(file)
                        ws[f'E{col_e}'].value = file
                        col_e += 1
            except:
                self.subf_dataset.append('N/E')
                ws[f'E{self.start_row}'].value = 'N/E'
            self.subf_dataset = []
            try:
                self.folder_st = os.listdir(f'{sub_folder}1. INFORMACIÓN GENERAL')
                col_f = self.start_row
                if len(os.listdir(f'{sub_folder}1. INFORMACIÓN GENERAL')) > 0:
                    print('\t1. INFORMACIÓN GENERAL')
                    for file in self.folder_st:
                        file = file.replace(ftitle,'').replace(' .pdf','.pdf').replace(' .png','.png').replace(' .jpg','.jpg').replace(' .jpeg','.jpeg').replace(' .webm','.webm').replace(' .mp4','.mp4')
                        print(f'\t\t>>> {file}')
                        self.subf_dataset.append(file)
                        self.collect_all_existing_files.append(file)
                        ws[f'F{col_f}'].value = file
                        col_f += 1
            except:
                self.subf_dataset.append('N/E')
                ws[f'F{self.start_row}'].value = 'N/E'
            self.subf_dataset = []
            try:
                self.folder_nd = os.listdir(f'{sub_folder}2. APROBACIONES CREDITICIAS')
                col_g = self.start_row
                if len(os.listdir(f'{sub_folder}2. APROBACIONES CREDITICIAS')) > 0:
                    print('\t2. APROBACIONES CREDITICIAS')
                    for file in self.folder_nd:
                        file = file.replace(ftitle,'').replace(' .pdf','.pdf').replace(' .png','.png').replace(' .jpg','.jpg').replace(' .jpeg','.jpeg').replace(' .webm','.webm').replace(' .mp4','.mp4')
                        print(f'\t\t>>> {file}')
                        self.subf_dataset.append(file)
                        self.collect_all_existing_files.append(file)
                        ws[f'G{col_g}'].value = file
                        col_g += 1
            except:
                self.subf_dataset.append('N/E')
                ws[f'G{self.start_row}'].value = 'N/E'
            self.subf_dataset = []
            try:
                self.folder_rd = os.listdir(f'{sub_folder}3. INFORMACIÓN PARA ANÁLISIS CAPACIDAD')
                col_h = self.start_row
                if len(os.listdir(f'{sub_folder}3. INFORMACIÓN PARA ANÁLISIS CAPACIDAD')) > 0:
                    print('\t3. INFORMACIÓN PARA ANÁLISIS CAPACIDAD')
                    for file in self.folder_rd:
                        file = file.replace(ftitle,'').replace(' .pdf','.pdf').replace(' .png','.png').replace(' .jpg','.jpg').replace(' .jpeg','.jpeg').replace(' .webm','.webm').replace(' .mp4','.mp4')
                        print(f'\t\t>>> {file}')
                        self.subf_dataset.append(file)
                        self.collect_all_existing_files.append(file)
                        ws[f'H{col_h}'].value = file
                        col_h += 1
            except:
                self.subf_dataset.append('N/E')
                ws[f'H{self.start_row}'].value = 'N/E'
            self.subf_dataset = []
            try:
                self.folder_th = os.listdir(f'{sub_folder}4. RESULTADOS DE ANÁLISIS')
                col_i = self.start_row
                if len(os.listdir(f'{sub_folder}4. RESULTADOS DE ANÁLISIS')) > 0:
                    print('\t4. RESULTADOS DE ANÁLISIS')
                    for file in self.folder_th:
                        file = file.replace(ftitle,'').replace(' .pdf','.pdf').replace(' .png','.png').replace(' .jpg','.jpg').replace(' .jpeg','.jpeg').replace(' .webm','.webm').replace(' .mp4','.mp4')
                        print(f'\t\t>>> {file}')
                        self.collect_all_existing_files.append(file)
                        self.subf_dataset.append(file)
                        ws[f'I{col_i}'].value = file
                        col_i += 1
            except:
                self.subf_dataset.append('N/E')
                ws[f'I{self.start_row}'].value = 'N/E'
            self.subf_dataset = []
            is_analysis = False
            is_cic = False
            is_cic_dated = False
            is_cicac = False
            is_contract = False
            is_affidavit = False
            is_id = False
            is_kyc = False
            is_document = False
            is_orden = False
            is_origin = False

            col_k = self.start_row
            for ff in self.collect_all_existing_files:
                ws[f'K{col_k}'].value = ff
                col_k += 1

            self.normalized_list = []
            for item in self.collect_all_existing_files:
                _item = item.lower()
                _item = _item.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('u','u')
                self.normalized_list.append(_item)
            for item in self.normalized_list:
                if item.__contains__('analisis'): is_analysis = True
                elif item.__contains__('cic'):
                    if item.__contains__('cicac'): is_cicac = True
                    else:
                        if item.__contains__('0') or item.__contains__('1') or item.__contains__('2') or item.__contains__('3') or item.__contains__('4') or item.__contains__('5') or item.__contains__('6') or item.__contains__('7') or item.__contains__('8') or item.__contains__('9'): is_cic_dated = True
                        else: is_cic = True
                elif item.__contains__('contrato'): is_contract = True
                elif item.__contains__('declarac'): is_affidavit = True
                elif item.__contains__('id'): is_id = True
                elif item.__contains__('kyc'): is_kyc = True
                elif item.__contains__('pagar') or  item.__contains__('letra'): is_document = True
                elif item.__contains__('orden'): is_orden = True
                elif item.__contains__('origen'): is_origin = True
            if is_analysis == False: self.subf_dataset.append('Análisis')
            if is_cic == False: self.subf_dataset.append('Autorización CIC')
            if is_cic_dated == False: self.subf_dataset.append('CIC')
            if is_cicac == False: self.subf_dataset.append('CICAC')
            if is_contract == False: self.subf_dataset.append('Contrato')
            if is_affidavit == False: self.subf_dataset.append('Declaración jurada')
            if is_id == False: self.subf_dataset.append('ID')
            if is_kyc == False: self.subf_dataset.append('KYC')
            if is_document == False: self.subf_dataset.append('Pagaré')
            if not is_orden and not is_origin: self.subf_dataset.append('Orden patronal')
            col_j = self.start_row
            print('\tDOCUMENTOS FALTANTES')
            print(self.subf_dataset)
            for file in self.subf_dataset:
                print(f'\t\t>>> {file}')
                ws[f'J{col_j}'].value = file
                col_j += 1
            ws.merge_cells(f'A{self.start_row}:A{ws.max_row}')
            ws[f'A{self.start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            counter = self.start_row
            row_rng = (ws.max_row - self.start_row) + 1
            try:
                for x in range(row_rng):
                    ws[f'B{counter}'].value = column_b
                    ws[f'C{counter}'].value = column_c
                    ws[f'D{counter}'].value = column_d
                    counter += 1
            except Exception as e: print(e)
            to_style = ws[f'A{self.start_row}:K{ws.max_row}']
            if self.highlight:
                self.highlight = False
                for ts in to_style:
                    for cell in ts:
                        cell.fill = self.intermitent_highlight
            else: self.highlight = True
        to_style = ws[f'A2:K{ws.max_row}']
        for ts in to_style:
            for cell in ts:
                cell.border = Border(top=self.wb_border, right=self.wb_border, bottom=self.wb_border, left=self.wb_border)
        wb.save(myexcel)